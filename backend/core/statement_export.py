"""Generate a company-specific cost sharing statement (Excel)."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from backend.core.translations import t, month_name

THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_FONT = Font(bold=True, size=11, color="FFFFFF")
RON_FMT = '#,##0.00 "RON"'


def generate_statement(filepath, company, result, month, year, monthly_input, lang="en", eur_rate=None):
    eur_rate = eur_rate or 5.1
    """Generate a single-company cost sharing statement.

    Args:
        filepath: output Excel path
        company: company dict
        result: single company allocation result dict
        month: int (1-12)
        year: int
        monthly_input: dict with all invoice/external values
        lang: "en" or "ro"
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    _titles = {"en": "Cost Sharing Statement", "ro": "Extras Costuri Comune", "tr": "Ortak Gider Ekstresi"}
    title = _titles.get(lang, _titles["en"])
    ws.title = title

    # Header
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="Premier Business Center")
    c.font = Font(bold=True, size=16)

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value=title).font = Font(bold=True, size=13)

    ws.merge_cells("A3:D3")
    mn = month_name(month, lang)
    period_label = f"{mn} {year}"
    ws.cell(row=3, column=1, value=period_label).font = Font(size=11, italic=True)

    row = 5

    # Company info
    company_label = t("excel_company", lang)
    ws.cell(row=row, column=1, value=company_label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=company["name"]).font = Font(bold=True, size=12)
    row += 1

    if company.get("building"):
        ws.cell(row=row, column=1, value=t("building", lang))
        ws.cell(row=row, column=2, value=company["building"])
        row += 1

    if company.get("office_no"):
        ws.cell(row=row, column=1, value=t("office_no", lang))
        ws.cell(row=row, column=2, value=company["office_no"])
        row += 1

    if company.get("contact_person"):
        ws.cell(row=row, column=1, value=t("contact_person", lang))
        ws.cell(row=row, column=2, value=company["contact_person"])
        row += 1

    row += 1

    # Expense breakdown table
    _cat_h = {"en": "Expense Category", "ro": "Categorie Cheltuiala", "tr": "Gider Kategorisi"}
    _amt_h = {"en": "Amount (RON)", "ro": "Suma (RON)", "tr": "Tutar (RON)"}
    headers = [
        _cat_h.get(lang, _cat_h["en"]),
        _amt_h.get(lang, _amt_h["en"]),
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = WHITE_FONT
        c.fill = BLUE_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN
    row += 1

    expenses = [
        (t("electricity", lang), result["electricity"]),
        (t("water", lang), result["water"]),
        (t("garbage", lang), result["garbage"]),
        (t("excel_gas_hotel", lang), result["gas_hotel"]),
        (t("excel_gas_gf", lang), result["gas_ground_floor"]),
        (t("excel_gas_ff", lang), result["gas_first_floor"]),
        (t("consumables", lang), result.get("consumables", 0)),
        (t("printer", lang), result.get("printer", 0)),
        (t("internet", lang), result.get("internet", 0)),
        (t("maintenance", lang) + (f" ({company.get('maintenance_rate_eur', 0):.2f} EUR \u00d7 {eur_rate})" if eur_rate and company.get("maintenance_rate_eur", 0) > 0 else ""), result.get("maintenance", 0)),
        (t("maintenance_vat", lang), result.get("maintenance_vat", 0)),
        (t("rent", lang) + (f" ({company.get('monthly_rent_eur', 0):.2f} EUR \u00d7 {eur_rate})" if eur_rate and company.get("monthly_rent_eur", 0) > 0 else ""), result.get("rent", 0)),
        (t("rent_vat", lang), result.get("rent_vat", 0)),
    ]

    for label, amount in expenses:
        if amount > 0:
            ws.cell(row=row, column=1, value=label).border = THIN
            c = ws.cell(row=row, column=2, value=amount)
            c.number_format = RON_FMT
            c.border = THIN
            c.alignment = Alignment(horizontal="right")
            row += 1

    # Total row
    total_label = t("excel_total", lang)
    ws.cell(row=row, column=1, value=total_label).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).border = THIN
    c = ws.cell(row=row, column=2, value=result["total"])
    c.font = Font(bold=True, size=12)
    c.number_format = RON_FMT
    c.border = THIN
    c.alignment = Alignment(horizontal="right")
    row += 2

    # Explanation
    _notes = {
        "en": (
            "This statement reflects your share of the shared building costs "
            "for the above period. Amounts are calculated based on your allocated "
            "area (m²) and number of persons, according to the cost sharing agreement."
        ),
        "ro": (
            "Acest extras reflecta cota dumneavoastra din costurile comune ale cladirii "
            "pentru perioada de mai sus. Sumele sunt calculate pe baza suprafetei alocate "
            "(m²) si a numarului de persoane, conform acordului de partajare a costurilor."
        ),
        "tr": (
            "Bu ekstre, yukarıdaki dönem için ortak bina giderlerinden size düşen payı "
            "yansıtmaktadır. Tutarlar, maliyet paylaşım sözleşmesine göre tahsis edilen "
            "alanınıza (m²) ve kişi sayınıza göre hesaplanmıştır."
        ),
    }
    note = _notes.get(lang, _notes["en"])

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(size=9, italic=True, color="666666")
    c.alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22

    wb.save(filepath)
    return filepath
