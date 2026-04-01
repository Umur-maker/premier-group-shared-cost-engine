import os
import tempfile
import openpyxl
import pytest
from excel_export import generate_excel


@pytest.fixture
def results():
    return [
        {"company_id": "comp-a", "company_name": "Company A",
         "electricity": 100.0, "water": 50.0, "garbage": 30.0,
         "gas_hotel": 0.0, "gas_ground_floor": 80.0, "gas_first_floor": 0.0, "total": 260.0},
        {"company_id": "hotel", "company_name": "Hotel",
         "electricity": 300.0, "water": 150.0, "garbage": 90.0,
         "gas_hotel": 500.0, "gas_ground_floor": 0.0, "gas_first_floor": 0.0, "total": 1040.0},
    ]


@pytest.fixture
def monthly_input():
    return {
        "electricity_total": 400.0, "garbage_total": 120.0, "water_total": 200.0,
        "hotel_gas_total": 500.0, "ground_floor_gas_total": 80.0, "first_floor_gas_total": 0.0,
        "external_electricity": 10.0, "external_water": 20.0, "external_garbage": 5.0,
        "external_hotel_gas": 0.0, "external_gf_gas": 0.0, "external_ff_gas": 0.0,
    }


@pytest.fixture
def ratios():
    return {
        "electricity": {"sqm_weight": 50, "headcount_weight": 50},
        "gas": {"sqm_weight": 80, "headcount_weight": 20},
        "water": {"sqm_weight": 30, "headcount_weight": 70},
        "garbage": {"sqm_weight": 30, "headcount_weight": 70},
    }


@pytest.fixture
def companies():
    return [
        {"id": "comp-a", "name": "Company A", "area_m2": 50.0, "headcount_default": 2,
         "floor": "ground_floor", "has_heating": True, "active": True,
         "electricity_eligible": True, "water_eligible": True, "garbage_eligible": True, "building": "C4"},
        {"id": "hotel", "name": "Hotel", "area_m2": 200.0, "headcount_default": 8,
         "floor": "hotel", "has_heating": True, "active": True,
         "electricity_eligible": True, "water_eligible": True, "garbage_eligible": True, "building": "C1"},
    ]


def _gen(results, monthly_input, ratios, companies, lang="en"):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    generate_excel(path, results, monthly_input, ratios, companies, lang)
    return path


def _all_cell_values(ws, max_row=60, max_col=8):
    """Collect all cell string values for searching."""
    vals = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                vals.append(str(v))
    return vals


def test_creates_file(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies)
    try:
        assert os.path.exists(path) and os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_three_sheets_english(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "en")
    try:
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Summary", "Detailed Breakdown", "Calculation Details"]
    finally:
        os.unlink(path)


def test_three_sheets_romanian(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "ro")
    try:
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Sumar", "Detalii pe Categorii", "Detalii Calcul"]
    finally:
        os.unlink(path)


def test_summary_data(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        assert ws.cell(row=2, column=1).value == "Company A"
        assert ws.cell(row=2, column=2).value == 260.0
    finally:
        os.unlink(path)


def test_no_elevator_in_calc_sheet(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Calculation Details"]
        vals = _all_cell_values(ws)
        text = " ".join(vals).lower()
        assert "elevator" not in text
    finally:
        os.unlink(path)


def test_person_label_not_headcount(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "en")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Calculation Details"]
        text = " ".join(_all_cell_values(ws))
        assert "Person" in text or "person" in text
        assert "Headcount" not in text
    finally:
        os.unlink(path)


# --- Currency formatting ---

def test_ron_currency_format_in_summary(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        fmt = ws.cell(row=2, column=2).number_format
        assert "RON" in fmt
    finally:
        os.unlink(path)


def test_ron_currency_format_in_detail(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Detailed Breakdown"]
        fmt = ws.cell(row=2, column=2).number_format
        assert "RON" in fmt
    finally:
        os.unlink(path)


# --- External usage section ---

def test_external_usage_section_exists(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "en")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Calculation Details"]
        vals = _all_cell_values(ws)
        text = " ".join(vals)
        assert "EXTERNAL USAGE" in text
        assert "ORIGINAL" in text
        assert "NET ALLOCABLE" in text
    finally:
        os.unlink(path)


def test_all_six_external_fields_in_excel(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "en")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Calculation Details"]
        text = " ".join(_all_cell_values(ws))
        assert "External Electricity" in text
        assert "External Water" in text
        assert "External Garbage" in text
        assert "External Hotel Gas" in text
        assert "External Ground Floor Gas" in text
        assert "External First Floor Gas" in text
    finally:
        os.unlink(path)


# --- Company numbering ---

def test_company_numbering_in_calc_sheet(results, monthly_input, ratios, companies):
    path = _gen(results, monthly_input, ratios, companies, "en")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Calculation Details"]
        vals = _all_cell_values(ws)
        assert "No." in vals
    finally:
        os.unlink(path)
