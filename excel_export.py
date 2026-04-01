import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from translations import t, floor_name

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
SECTION_FONT = Font(bold=True, size=12)
SUBSECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_cell(ws, row, col, value, is_currency=False, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    if is_currency:
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")
    if bold:
        cell.font = Font(bold=True)
    return cell


def _write_summary_sheet(wb, results, lang):
    ws = wb.create_sheet(t("excel_summary", lang))
    headers = [t("excel_company", lang), t("excel_total_payment", lang)]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(results, 2):
        _write_cell(ws, i, 1, r["company_name"])
        _write_cell(ws, i, 2, r["total"], is_currency=True)

    total_row = len(results) + 2
    _write_cell(ws, total_row, 1, t("excel_total", lang), bold=True)
    _write_cell(ws, total_row, 2, round(sum(r["total"] for r in results), 2), is_currency=True, bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22


def _write_detailed_sheet(wb, results, lang):
    ws = wb.create_sheet(t("excel_detail", lang))
    headers = [t("excel_company", lang), t("electricity", lang), t("water", lang),
               t("garbage", lang), t("excel_gas_hotel", lang),
               t("excel_gas_gf", lang), t("excel_gas_ff", lang), t("excel_total", lang)]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    keys = ["electricity", "water", "garbage",
            "gas_hotel", "gas_ground_floor", "gas_first_floor", "total"]

    for i, r in enumerate(results, 2):
        _write_cell(ws, i, 1, r["company_name"])
        for j, key in enumerate(keys, 2):
            _write_cell(ws, i, j, r[key], is_currency=True)

    total_row = len(results) + 2
    _write_cell(ws, total_row, 1, t("excel_total", lang), bold=True)
    for j, key in enumerate(keys, 2):
        _write_cell(ws, total_row, j, round(sum(r[key] for r in results), 2), is_currency=True, bold=True)

    ws.column_dimensions["A"].width = 25
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_calculation_sheet(wb, results, monthly_input, ratios, companies, lang):
    ws = wb.create_sheet(t("excel_calc", lang))
    row = 1

    # Section 1: Input Values
    ws.cell(row=row, column=1, value=t("excel_input_values", lang)).font = SECTION_FONT
    row += 1

    bill_labels = [
        (t("electricity", lang), "electricity_total"),
        (t("water", lang), "water_total"),
        (t("garbage", lang), "garbage_total"),
        (t("hotel_gas", lang), "hotel_gas_total"),
        (t("ground_floor_gas", lang), "ground_floor_gas_total"),
        (t("first_floor_gas", lang), "first_floor_gas_total"),
    ]
    for label, key in bill_labels:
        ws.cell(row=row, column=1, value=f"{label} (RON)")
        ws.cell(row=row, column=2, value=monthly_input[key]).number_format = "#,##0.00"
        row += 1

    row += 1

    # External usage
    ext_labels = [
        (t("excel_external_electricity", lang), "external_electricity"),
        (t("excel_external_water", lang), "external_water"),
        (t("excel_external_garbage", lang), "external_garbage"),
        (t("excel_external_hotel_gas", lang), "external_hotel_gas"),
        (t("excel_external_gf_gas", lang), "external_gf_gas"),
        (t("excel_external_ff_gas", lang), "external_ff_gas"),
    ]
    for label, key in ext_labels:
        val = monthly_input.get(key, 0)
        if val > 0:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val).number_format = "#,##0.00"
            row += 1

    row += 1

    # Section 2: Net Allocable
    ws.cell(row=row, column=1, value=t("excel_net_amounts", lang)).font = SECTION_FONT
    row += 1

    net_pairs = [
        ("electricity", "electricity_total", "external_electricity"),
        ("water", "water_total", "external_water"),
        ("garbage", "garbage_total", "external_garbage"),
        ("hotel_gas", "hotel_gas_total", "external_hotel_gas"),
        ("ground_floor_gas", "ground_floor_gas_total", "external_gf_gas"),
        ("first_floor_gas", "first_floor_gas_total", "external_ff_gas"),
    ]
    for type_key, total_key, ext_key in net_pairs:
        ext = monthly_input.get(ext_key, 0)
        net = monthly_input[total_key] - ext
        label_type = t(type_key, lang) if type_key in ("electricity", "water", "garbage") else t(type_key, lang)
        if ext > 0:
            label = t("excel_after_external", lang, type=label_type)
        else:
            label = label_type
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=round(net, 2)).number_format = "#,##0.00"
        row += 1

    row += 1

    # Section 3: Ratios
    ws.cell(row=row, column=1, value=t("excel_ratios", lang)).font = SECTION_FONT
    row += 1
    for expense_type, weights in ratios.items():
        label = expense_type.capitalize()
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=t("excel_ratio_format", lang,
                                            sqm=weights["sqm_weight"], hc=weights["headcount_weight"]))
        row += 1

    row += 1

    # Section 4: Company Data
    active = [c for c in companies if c["active"]]
    ws.cell(row=row, column=1, value=t("excel_company_data", lang)).font = SECTION_FONT
    row += 1

    headers = [t("excel_company", lang), t("area_m2", lang), t("excel_persons", lang),
               t("floor", lang), t("excel_has_heating", lang),
               t("excel_sqm_pct", lang), t("excel_person_pct", lang)]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))
    row += 1

    total_sqm = sum(c["area_m2"] for c in active)
    total_hc = sum(c["headcount_default"] for c in active)

    for c in active:
        ws.cell(row=row, column=1, value=c["name"])
        ws.cell(row=row, column=2, value=c["area_m2"])
        ws.cell(row=row, column=3, value=c["headcount_default"])
        ws.cell(row=row, column=4, value=floor_name(c["floor"], lang))
        ws.cell(row=row, column=5, value=t("excel_yes", lang) if c["has_heating"] else t("excel_no", lang))
        sqm_pct = c["area_m2"] / total_sqm if total_sqm else 0
        hc_pct = c["headcount_default"] / total_hc if total_hc else 0
        ws.cell(row=row, column=6, value=round(sqm_pct * 100, 2))
        ws.cell(row=row, column=6).number_format = '0.00"%"'
        ws.cell(row=row, column=7, value=round(hc_pct * 100, 2))
        ws.cell(row=row, column=7).number_format = '0.00"%"'
        row += 1

    ws.cell(row=row, column=1, value=t("excel_total", lang)).font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_sqm).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_hc).font = Font(bold=True)
    row += 2

    # Section 5: Eligible Groups
    ws.cell(row=row, column=1, value=t("excel_eligible", lang)).font = SECTION_FONT
    row += 1

    expense_groups = [
        (t("electricity", lang), [c for c in active if c["electricity_eligible"]]),
        (t("water", lang), [c for c in active if c["water_eligible"]]),
        (t("garbage", lang), [c for c in active if c["garbage_eligible"]]),
        (t("excel_gas_hotel", lang), [c for c in active if c["floor"] == "hotel" and c["has_heating"]]),
        (t("excel_gas_gf", lang), [c for c in active if c["floor"] == "ground_floor" and c["has_heating"]]),
        (t("excel_gas_ff", lang), [c for c in active if c["floor"] == "first_floor" and c["has_heating"]]),
    ]

    for expense_name, eligible in expense_groups:
        ws.cell(row=row, column=1, value=expense_name).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = SUBSECTION_FILL
        eligible_sqm = sum(c["area_m2"] for c in eligible)
        eligible_hc = sum(c["headcount_default"] for c in eligible)
        ws.cell(row=row, column=2, value=t("excel_n_companies", lang, n=len(eligible))).fill = SUBSECTION_FILL
        ws.cell(row=row, column=3, value=t("excel_total_sqm", lang, v=eligible_sqm)).fill = SUBSECTION_FILL
        ws.cell(row=row, column=4, value=t("excel_total_persons", lang, v=eligible_hc)).fill = SUBSECTION_FILL
        row += 1

        for c in eligible:
            hc = c["headcount_default"]
            sqm_ratio = c["area_m2"] / eligible_sqm if eligible_sqm > 0 else 0
            hc_ratio = hc / eligible_hc if eligible_hc > 0 else 0
            ws.cell(row=row, column=1, value=f"  {c['name']}")
            ws.cell(row=row, column=2, value=f"{c['area_m2']} m\u00b2 ({sqm_ratio*100:.2f}%)")
            ws.cell(row=row, column=3, value=f"{hc} ({hc_ratio*100:.2f}%)")
            result_row = next((r for r in results if r["company_id"] == c["id"]), None)
            if result_row:
                key_map = {
                    t("electricity", lang): "electricity",
                    t("water", lang): "water",
                    t("garbage", lang): "garbage",
                    t("excel_gas_hotel", lang): "gas_hotel",
                    t("excel_gas_gf", lang): "gas_ground_floor",
                    t("excel_gas_ff", lang): "gas_first_floor",
                }
                key = key_map.get(expense_name)
                if key:
                    ws.cell(row=row, column=4, value=result_row[key])
                    ws.cell(row=row, column=4).number_format = "#,##0.00"
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 35
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22


def generate_excel(filepath, results, monthly_input, ratios, companies, lang="en"):
    """Generate a 3-sheet Excel workbook with cost allocation results."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, results, lang)
    _write_detailed_sheet(wb, results, lang)
    _write_calculation_sheet(wb, results, monthly_input, ratios, companies, lang)

    wb.save(filepath)
    return filepath
